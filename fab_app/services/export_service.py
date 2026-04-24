from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fab_app.models.dashboard import normalize_page
from fab_app.services.dashboard_service import get_filtered_export_rows


EXPORT_COLUMNS = {
    "angle": [
        {"key": "source", "header": "Source"},
        {"key": "category", "header": "Angle"},
        {"key": "project", "header": "Project"},
        {"key": "vendor", "header": "Vendor"},
        {"key": "aaMC", "header": "AA MC"},
        {"key": "stn", "header": "STN"},
        {"key": "date", "header": "Date"},
        {"key": "metricValue", "header": "OOS Rate", "number_format": "0.00%"},
        {"key": "failCount", "header": "Fail"},
        {"key": "inputCount", "header": "Input"},
    ],
    "oc": [
        {"key": "source", "header": "Source"},
        {"key": "category", "header": "Series"},
        {"key": "project", "header": "Project"},
        {"key": "aaMC", "header": "AA MC"},
        {"key": "stn", "header": "STN"},
        {"key": "date", "header": "Date"},
        {"key": "metricValue", "header": "OC Value", "number_format": "0.000"},
        {"key": "severityValue", "header": "|OC|", "number_format": "0.000"},
        {"key": "inputCount", "header": "Input"},
    ],
    "lens": [
        {"key": "source", "header": "Source"},
        {"key": "category", "header": "Lens PP"},
        {"key": "config", "header": "Config"},
        {"key": "project", "header": "Project"},
        {"key": "aaMC", "header": "AA MC"},
        {"key": "aaTime", "header": "AA Time"},
        {"key": "date", "header": "Test Time"},
        {"key": "metricValue", "header": "Fail Rate", "number_format": "0.00%"},
        {"key": "failCount", "header": "Fail"},
        {"key": "inputCount", "header": "Input"},
    ],
}

_COLUMN_MAPS = {
    page: {column["key"]: column for column in columns}
    for page, columns in EXPORT_COLUMNS.items()
}


def export_excel_file(page, filters, columns, refresh=False):
    normalized_page = normalize_page(page)
    selected_columns = _resolve_columns(columns or [], _COLUMN_MAPS[normalized_page])
    rows = get_filtered_export_rows(normalized_page, filters or {}, refresh=refresh)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{normalized_page.title()} Export"

    for column_index, column in enumerate(selected_columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column["header"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="0F4C81")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(selected_columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=row.get(column["key"]))
            if "number_format" in column:
                cell.number_format = column["number_format"]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(selected_columns))}{max(len(rows) + 1, 1)}"
    _autofit_columns(worksheet, selected_columns, rows)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"{normalized_page}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer, filename


def _resolve_columns(columns, column_map):
    if not columns:
        raise ValueError("Please choose at least one export column.")

    selected_columns = []
    seen_keys = set()
    for key in columns:
        normalized_key = str(key or "").strip()
        if normalized_key in column_map and normalized_key not in seen_keys:
            selected_columns.append(column_map[normalized_key])
            seen_keys.add(normalized_key)

    if not selected_columns:
        raise ValueError("Please choose at least one export column.")

    return selected_columns


def _autofit_columns(worksheet, selected_columns, rows):
    for column_index, column in enumerate(selected_columns, start=1):
        max_length = len(column["header"])
        for row in rows:
            value = row.get(column["key"])
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 4, 28)
